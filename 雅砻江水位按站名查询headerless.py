import os
import time

import openpyxl
# 1. import the necessary packages
import pandas as pd
from bs4 import BeautifulSoup
from openpyxl.utils.dataframe import dataframe_to_rows
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.wait import WebDriverWait

# 2. set the chrome driver
options = webdriver.ChromeOptions()
options.add_argument('--headless')
options.add_argument('--no-sandbox')
options.add_argument('--disable-dev-shm-usage')
options.add_argument('/usr/local/bin/chromedriver')

# 3. open the website
# 设置引擎为Chrome，从本地打开一个Chrome浏览器 in headless mode
driver = webdriver.Chrome(options=options)

# 打开网页
driver.get('https://tftb.sczwfw.gov.cn:8085/hos-server/pub/jmas/jmasbucket/jmopen_files/unzip/6e5032129863494a94bb2e2e7a2e9748/sltqszdsksssqxxpc/index.html#/')

# 等待页面加载完成
wait = WebDriverWait(driver, 15)

# 找到搜索栏
# search_box = driver.find_element(By.CSS_SELECTOR, '#app > div > div.centerbody > div:nth-child(1) > div > form > div:nth-child(3) > div > div > input')
search_box = wait.until(EC.visibility_of_element_located((By.CSS_SELECTOR, "#app > div > div.centerbody > div:nth-child(1) > div > form > div:nth-child(3) > div > div > input")))

# 在搜索框中输入关键字并搜索
search_box.send_keys('雅砻江')
time.sleep(10)

# 查找搜索按钮并点击
search_button = driver.find_element(By.CSS_SELECTOR, '#app > div > div.centerbody > div:nth-child(1) > div > button.el-button.blue_button.el-button--default.el-button--mini.is-plain')

# wait for the page to load

driver.execute_script("arguments[0].click();", search_button)

time.sleep(10)

# Parse the HTML content
soup = BeautifulSoup(driver.page_source, 'html.parser')

# Find the table header by class name
# table_header = soup.find('table', {'class': 'el-table__header'}) 

# find the table content by class name and loop over the rows in the table and extract the data you need 
table = soup.find('table', {'class': 'el-table__body'})
data = []
for row in table.find_all('tr'): # type: ignore
        # Find the columns in the row
        cols = row.find_all('td')
        # Extract the text from the columns
        cols = [ele.text.strip() for ele in cols]
        # append data
        data.append([ele for ele in cols if ele])

# Insert the table header to the data
# data.insert(0, [ele.text.strip() for ele in table_header.find_all('th')]) # type: ignore
  
# Convert the data to a DataFrame
df = pd.DataFrame(data)

# Create a dictionary to store the data for each station
station_data = {}

# Loop over the rows in the DataFrame and store the data for each station in the array
for index, row in df.iterrows():
    station_name = row[0]
    if station_name not in station_data:
       station_data[station_name] = []
    station_data[station_name].append(row)

# Create a new Excel workbook
if os.path.exists('雅砻江站名水位-PC.xlsx'):
    workbook = openpyxl.load_workbook('雅砻江站名水位-PC.xlsx')
else:
    workbook = openpyxl.Workbook()

# Loop over the stations in the dictionary and create a new sheet for each station
for station_name, station_rows in station_data.items():
    if station_name in workbook.sheetnames:
        sheet = workbook[station_name]
        existing_data = pd.read_excel('雅砻江站名水位-PC.xlsx', sheet_name=station_name, header=None)
    else:
        sheet = workbook.create_sheet(station_name)
        existing_data = pd.DataFrame()
    
    # Convert station_rows to dataframe
    new_data = pd.DataFrame(station_rows)
      
    # Append new data to existing data
    updated_data = existing_data.append(new_data) # type: ignore
    
    # Drop duplicates based on column 10 except the header
    updated_data = updated_data.drop_duplicates(subset=[updated_data.columns[9]], keep='first', ignore_index=True)
    
    # Sort the data by column 10 except the header
    updated_data = updated_data.iloc[1:].sort_values(by=[9], axis=0, ascending=True, inplace=False, kind='quicksort', na_position='last') # type: ignore
    
    # Delete existing data in sheet
    sheet.delete_rows(2, sheet.max_row)     
    
    # Write updated data to sheet
    for r in dataframe_to_rows(updated_data, index=False, header=False):
        sheet.append(r)
   
    
# Save the Excel workbook
workbook.save('雅砻江站名水位-PC.xlsx')

driver.close() # 关闭浏览器